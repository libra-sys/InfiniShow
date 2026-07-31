"""文件服务."""

import io
import os
import tempfile
from datetime import datetime
from typing import Any

from sqlalchemy import desc, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.constants import ALLOWED_UPLOAD_EXTENSIONS, FileType, MAX_UPLOAD_SIZE_MB
from app.core.exceptions import BusinessException, NotFoundException
from app.models.file_record import FileRecord
from app.models.user import User


async def upload_file(db: AsyncSession, user: User, file_name: str, file_content: bytes) -> FileRecord:
    """上传文件到 MinIO 并创建记录."""
    # 校验文件大小
    max_bytes = MAX_UPLOAD_SIZE_MB * 1024 * 1024
    if len(file_content) > max_bytes:
        raise BusinessException(f"文件大小超过 {MAX_UPLOAD_SIZE_MB}MB 限制")

    # 校验扩展名
    _, ext = os.path.splitext(file_name)
    if ext.lower() not in ALLOWED_UPLOAD_EXTENSIONS:
        raise BusinessException(f"不支持的文件格式: {ext}")

    # 判断文件类型
    file_type = FileType.CSV.value if ext.lower() == ".csv" else FileType.EXCEL.value

    # 生成存储键
    import ulid
    storage_key = f"uploads/{user.id}/{str(ulid.new())}/{file_name}"

    # 上传到 MinIO
    from app.services.minio_client import minio_client

    minio_client.upload_file(storage_key, file_content)

    # 解析文件获取列信息
    columns, row_count = _parse_file(file_content, ext)

    file_record = FileRecord(
        user_id=user.id,
        original_name=file_name,
        file_type=file_type,
        storage_key=storage_key,
        size_bytes=len(file_content),
        columns=columns,
        row_count=row_count,
    )
    db.add(file_record)
    await db.commit()
    await db.refresh(file_record)

    return file_record


def _parse_file(file_content: bytes, ext: str) -> tuple[list[dict[str, Any]] | None, int | None]:
    """解析文件获取列信息."""
    try:
        if ext.lower() == ".csv":
            import csv
            text = file_content.decode("utf-8-sig")
            reader = csv.reader(io.StringIO(text))
            headers = next(reader, None)
            row_count = sum(1 for _ in reader)
            columns = [{"name": h, "type": "string", "sample": ""} for h in headers] if headers else None
            return columns, row_count
        else:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(file_content), read_only=True)
            ws = wb.active
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            row_count = ws.max_row - 1 if ws.max_row else 0
            columns = [{"name": str(h), "type": "string", "sample": ""} for h in headers if h]
            wb.close()
            return columns, row_count
    except Exception:
        return None, None


async def list_files(db: AsyncSession, user: User) -> list[FileRecord]:
    """获取用户的文件列表."""
    result = await db.execute(
        select(FileRecord)
        .where(FileRecord.user_id == user.id)
        .order_by(desc(FileRecord.created_at))
    )
    return list(result.scalars().all())


async def download_file(db: AsyncSession, file_id: str, user: User) -> tuple[str, bytes]:
    """下载文件."""
    file_record = await db.get(FileRecord, file_id)
    if not file_record or file_record.user_id != user.id:
        raise NotFoundException("文件不存在")

    from app.services.minio_client import minio_client
    content = minio_client.download_file(file_record.storage_key)
    return file_record.original_name, content


async def delete_file(db: AsyncSession, file_id: str, user: User) -> None:
    """删除文件."""
    file_record = await db.get(FileRecord, file_id)
    if not file_record or file_record.user_id != user.id:
        raise NotFoundException("文件不存在")

    from app.services.minio_client import minio_client
    minio_client.delete_file(file_record.storage_key)

    await db.delete(file_record)
    await db.commit()
