"""提示词工程模块 — 放大产品和 AI 能力的核心.

构建结构化系统提示词，引导 InfiniSynapse Agent 输出标准化的经营分析报告。
包含：角色设定、任务约束、输出格式规范、溯源要求、场景上下文。
"""

import json
import logging
from typing import Any

logger = logging.getLogger("infinshow.prompt")

# ─── 系统角色与能力设定 ───
SYSTEM_ROLE = """你是一位资深经营分析专家，拥有以下能力：
1. **多维度健康度评估**：从盈利能力、运营效率、客户满意、成本管控、成长潜力五个维度量化评分（0-100）
2. **可信溯源分析**：每个结论必须标注数据来源行号、计算公式、验证过程，确保全链路可溯源
3. **行业基准对比**：基于行业平均水平给出 percentile 排名
4. **行动建议生成**：按优先级（高/中/低）输出可落地的改善建议，含预期效果
5. **异常检测**：自动识别数据中的异常值、趋势突变、潜在风险"""

# ─── 输出格式规范 ───
OUTPUT_FORMAT = """
请严格按以下 JSON 结构输出分析结果（用 ```json 包裹）：

```json
{
  "overall_score": "评级字母（A+/A/B+/B/C+/C/D）",
  "health_scores": [
    {"dimension": "盈利能力", "score": 0-100, "weight": 0.25, "benchmark": "行业均值75", "percentile": 0-100},
    {"dimension": "运营效率", "score": 0-100, "weight": 0.20, "benchmark": "行业均值70", "percentile": 0-100},
    {"dimension": "客户满意", "score": 0-100, "weight": 0.20, "benchmark": "行业均值80", "percentile": 0-100},
    {"dimension": "成本管控", "score": 0-100, "weight": 0.20, "benchmark": "行业均值65", "percentile": 0-100},
    {"dimension": "成长潜力", "score": 0-100, "weight": 0.15, "benchmark": "行业均值60", "percentile": 0-100}
  ],
  "kpis": [
    {"name": "指标名称", "value": "数值", "unit": "单位", "trend": "up/down/flat", "yoy": "同比变化", "benchmark": "行业基准"}
  ],
  "conclusions": [
    {
      "metric": "指标名",
      "value": "计算结果",
      "level": "consistent/questionable/inconsistent",
      "formula": "计算公式（如 SUM(D2:D100)）",
      "source_rows": [行号列表],
      "verification_process": "核对过程描述",
      "raw_data_sample": "关联原始数据片段"
    }
  ],
  "actions": [
    {
      "title": "建议标题",
      "priority": "high/medium/low",
      "description": "详细说明",
      "expected_effect": "预期效果",
      "timeline": "建议执行周期"
    }
  ],
  "charts": [
    {"id": "pie_profit", "type": "pie", "title": "品类利润占比", "data": [{"name": "品类A", "value": 100}]}
  ],
  "risk_warnings": [
    {"level": "high/medium/low", "description": "风险描述", "recommendation": "应对建议"}
  ],
  "summary": "一段话总结经营状况（50-100字）"
}
```"""

# ─── 溯源与可信度约束 ───
TRACEABILITY_RULES = """
**溯源约束（硬性要求）：**
1. 每个结论必须关联到具体的输入数据行号（source_rows）
2. 必须给出计算公式（formula），使用 Excel 公式语法
3. level 字段标注可信度：consistent（数据一致）/ questionable（存疑）/ inconsistent（不符）
4. 如果数据不足以支撑结论，标注 level=questionable 并说明原因
5. 禁止编造未在输入数据中出现的数据"""

# ─── 场景上下文模板 ───
SCENARIO_CONTEXT = {
    "S01": "外卖餐饮行业，关注订单量、客单价、差评率、出餐时长、平台扣点、食材成本率",
    "S02": "电商行业，关注转化率、退货率、广告ROI、GMV、客单价",
    "S03": "便利店行业，关注坪效、库存周转、缺货率、损耗率",
    "S04": "生鲜行业，关注损耗率、毛利率、复购率、会员占比",
    "S05": "美业行业，关注客单价、翻台率、会员占比、预约等待",
    "S06": "教培行业，关注消课率、续费率、退费率、获客成本",
    "S07": "健身行业，关注会员留存、课消率、续卡率",
    "S08": "宠物服务行业，关注服务占比、复购率、客单价",
    "S09": "汽车后市场，关注单车产值、回厂率、毛利",
    "S10": "母婴零售，关注会员销售占比、周转天数、连带率",
    "S11": "文创手作，关注体验课转化、客单价、复购率",
    "S12": "综合零售，关注客单价、品类贡献、库存周转",
}


def build_system_prompt(
    scenario_code: str,
    user_inputs: dict[str, Any] | None = None,
    file_summaries: list[dict[str, Any]] | None = None,
) -> str:
    """构建结构化系统提示词.

    Args:
        scenario_code: 场景代码（S01-S12）
        user_inputs: 用户表单输入的键值对
        file_summaries: 用户上传文件的摘要信息 [{name, columns, row_count, sample_rows}]

    Returns:
        结构化的提示词文本
    """
    scenario_context = SCENARIO_CONTEXT.get(scenario_code, "通用经营分析")
    parts: list[str] = []

    # 1. 角色设定
    parts.append(f"# 角色设定\n{SYSTEM_ROLE}")

    # 2. 场景上下文
    parts.append(f"# 分析场景\n当前分析场景：{scenario_code}\n行业背景：{scenario_context}")

    # 3. 用户输入数据
    if user_inputs:
        formatted_inputs = json.dumps(user_inputs, ensure_ascii=False, indent=2)
        parts.append(f"# 用户输入的经营数据\n```json\n{formatted_inputs}\n```")

    # 4. 文件数据摘要
    if file_summaries:
        file_desc_parts = []
        for i, f in enumerate(file_summaries, 1):
            cols = ", ".join([c.get("name", "") for c in (f.get("columns") or [])[:20]])
            file_desc_parts.append(
                f"文件{i}: {f.get('name', '未命名')} "
                f"(类型: {f.get('file_type', '未知')}, "
                f"行数: {f.get('row_count', '未知')}, "
                f"列: [{cols}])"
            )
            # 附带前5行样本数据
            if f.get("sample_rows"):
                sample = json.dumps(f["sample_rows"][:5], ensure_ascii=False, indent=2)
                file_desc_parts.append(f"  样本数据（前5行）：\n{sample}")
        parts.append("# 上传文件数据\n" + "\n".join(file_desc_parts))

    # 5. 溯源约束
    parts.append(TRACEABILITY_RULES)

    # 6. 输出格式
    parts.append(OUTPUT_FORMAT)

    # 7. 行动指引
    parts.append(
        "# 行动指引\n"
        "1. 优先分析用户输入的核心指标，计算衍生指标\n"
        "2. 对每个指标给出行业基准对比和百分位排名\n"
        "3. 识别3-5个关键风险点并给出应对建议\n"
        "4. 确保所有结论可溯源到输入数据\n"
        "5. 如果上传了文件，优先从文件数据中提取指标\n"
        "6. 如果数据不足，明确标注 level=questionable"
    )

    return "\n\n".join(parts)


def build_followup_prompt(question: str, report_context: dict[str, Any] | None = None) -> str:
    """构建追问提示词.

    Args:
        question: 用户追问问题
        report_context: 已有报告上下文（用于多轮对话）

    Returns:
        追问提示词
    """
    parts = [f"用户追问：{question}"]

    if report_context:
        parts.append(
            f"已有分析结果摘要：评级 {report_context.get('overall_score', 'N/A')}，"
            f"健康度维度：{json.dumps(report_context.get('health_scores', []), ensure_ascii=False)[:200]}"
        )

    parts.append(
        "请基于已有分析结果回答用户追问。"
        "如果追问涉及新指标，请给出计算公式和数据来源行号。"
        "如果是行动建议追问，请给出具体可执行的步骤和预期效果。"
    )

    return "\n\n".join(parts)


def build_file_summary_prompt(file_content: bytes, file_name: str) -> dict[str, Any]:
    """从文件内容提取摘要，用于注入 prompt 上下文.

    Args:
        file_content: 文件二进制内容
        file_name: 文件名

    Returns:
        {name, file_type, columns, row_count, sample_rows}
    """
    import io
    import os

    _, ext = os.path.splitext(file_name)
    ext = ext.lower()

    summary: dict[str, Any] = {
        "name": file_name,
        "file_type": "csv" if ext == ".csv" else "excel",
        "columns": [],
        "row_count": 0,
        "sample_rows": [],
    }

    try:
        if ext == ".csv":
            import csv

            text = file_content.decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(text))
            summary["columns"] = [{"name": h, "type": "string"} for h in reader.fieldnames or []]

            rows = []
            for i, row in enumerate(reader):
                if i >= 5:
                    break
                rows.append(row)
            summary["sample_rows"] = rows

            # 统计总行数
            reader2 = csv.reader(io.StringIO(text))
            summary["row_count"] = max(0, sum(1 for _ in reader2) - 1)

        elif ext in (".xlsx", ".xls"):
            from openpyxl import load_workbook

            wb = load_workbook(io.BytesIO(file_content), read_only=True)
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            headers = next(rows_iter, None)
            if headers:
                summary["columns"] = [{"name": str(h), "type": "string"} for h in headers if h]

            sample = []
            for i, row in enumerate(rows_iter):
                if i >= 5:
                    break
                if headers:
                    sample.append({str(headers[j]): str(row[j]) if row[j] is not None else "" for j in range(min(len(headers), len(row)))})
                else:
                    sample.append(list(row))
            summary["sample_rows"] = sample
            summary["row_count"] = ws.max_row - 1 if ws.max_row else 0
            wb.close()
    except Exception as e:
        logger.warning("File summary extraction failed for %s: %s", file_name, e)

    return summary
