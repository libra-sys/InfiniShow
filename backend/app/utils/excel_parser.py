"""Excel/CSV 解析与字段识别."""

import csv
import io
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.core.constants import ALLOWED_UPLOAD_EXTENSIONS, MAX_UPLOAD_SIZE_MB
from app.core.exceptions import BusinessException


def validate_upload_file(filename: str, size_bytes: int) -> None:
    """校验上传文件."""
    ext = Path(filename).suffix.lower()
    if ext not in ALLOWED_UPLOAD_EXTENSIONS:
        raise BusinessException(f"不支持的文件格式，仅支持: {', '.join(ALLOWED_UPLOAD_EXTENSIONS)}")
    if size_bytes > MAX_UPLOAD_SIZE_MB * 1024 * 1024:
        raise BusinessException(f"文件大小超过限制，最大 {MAX_UPLOAD_SIZE_MB}MB")


def parse_excel(file_bytes: bytes) -> tuple[list[dict[str, Any]], list[dict[str, str]], int]:
    """解析 Excel 文件.

    Returns:
        rows: 数据行列表
        columns: 列信息列表 [{name, type, sample}]
        row_count: 数据行数
    """
    workbook = load_workbook(io.BytesIO(file_bytes), data_only=True)
    sheet = workbook.active
    if sheet is None:
        raise BusinessException("Excel 文件没有工作表")

    headers = [str(cell.value) if cell.value is not None else f"column_{i}" for i, cell in enumerate(sheet[1])]
    rows: list[dict[str, Any]] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = {}
        for i, header in enumerate(headers):
            row_data[header] = row[i] if i < len(row) else None
        rows.append(row_data)

    columns = infer_columns(headers, rows)
    return rows, columns, len(rows)


def parse_csv(file_bytes: bytes, encoding: str = "utf-8") -> tuple[list[dict[str, Any]], list[dict[str, str]], int]:
    """解析 CSV 文件."""
    try:
        text = file_bytes.decode(encoding)
    except UnicodeDecodeError:
        text = file_bytes.decode("gbk")

    reader = csv.DictReader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        raise BusinessException("CSV 文件为空或没有表头")

    headers = list(rows[0].keys())
    columns = infer_columns(headers, rows)
    return rows, columns, len(rows)


def infer_columns(headers: list[str], rows: list[dict[str, Any]]) -> list[dict[str, str]]:
    """推断列信息."""
    columns = []
    for header in headers:
        sample_values = [row.get(header) for row in rows[:5] if row.get(header) is not None]
        column_type = infer_column_type(sample_values)
        columns.append({
            "name": header,
            "type": column_type,
            "sample": str(sample_values[0]) if sample_values else "",
        })
    return columns


def infer_column_type(values: list[Any]) -> str:
    """推断列数据类型."""
    if not values:
        return "string"

    numeric_count = 0
    date_count = 0
    for value in values:
        if value is None:
            continue
        str_value = str(value).strip()
        if not str_value:
            continue
        try:
            float(str_value.replace(",", ""))
            numeric_count += 1
            continue
        except ValueError:
            pass
        if "/" in str_value or "-" in str_value and len(str_value) >= 8:
            date_count += 1
            continue

    total = len(values)
    if numeric_count / total > 0.5:
        return "number"
    if date_count / total > 0.5:
        return "date"
    return "string"


def parse_upload_file(filename: str, file_bytes: bytes) -> tuple[list[dict[str, Any]], list[dict[str, str]], int]:
    """解析上传文件（Excel/CSV）."""
    ext = Path(filename).suffix.lower()
    if ext in (".xlsx", ".xls"):
        return parse_excel(file_bytes)
    if ext == ".csv":
        return parse_csv(file_bytes)
    raise BusinessException(f"不支持的文件格式: {ext}")
